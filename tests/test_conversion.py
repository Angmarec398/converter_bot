import csv
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from openpyxl import Workbook

from handlers.file import _convert_xlsx_to_csv


def test_convert_xlsx_to_csv():
    """Проверяет корректную конвертацию xlsx файла в csv."""
    with TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        
        # Создаем тестовый xlsx файл
        xlsx_path = tmp_dir_path / "test_data.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        
        # Заполняем данными
        sheet["A1"] = "Имя"
        sheet["B1"] = "Фамилия"
        sheet["C1"] = "Возраст"
        sheet["A2"] = "Иван"
        sheet["B2"] = "Иванов"
        sheet["C2"] = 25
        sheet["A3"] = "Мария"
        sheet["B3"] = "Петрова"
        sheet["C3"] = 30
        
        workbook.save(xlsx_path)
        
        # Конвертируем в csv
        csv_path = tmp_dir_path / "test_data.csv"
        _convert_xlsx_to_csv(xlsx_path, csv_path)
        
        # Проверяем, что csv файл создан
        assert csv_path.exists(), "CSV файл должен быть создан"
        
        # Проверяем содержимое csv файла
        with csv_path.open("r", encoding="cp1251") as f:
            reader = csv.reader(f, delimiter=",")
            rows = list(reader)
            
            # Проверяем количество строк
            assert len(rows) == 3, "Должно быть 3 строки"
            
            # Проверяем заголовок
            assert rows[0] == ["Имя", "Фамилия", "Возраст"], "Заголовок должен совпадать"
            
            # Проверяем первую строку данных
            assert rows[1] == ["Иван", "Иванов", "25"], "Первая строка данных должна совпадать"
            
            # Проверяем вторую строку данных
            assert rows[2] == ["Мария", "Петрова", "30"], "Вторая строка данных должна совпадать"


def test_convert_xlsx_to_csv_with_empty_cells():
    """Проверяет обработку пустых ячеек в xlsx файле."""
    with TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        
        # Создаем xlsx файл с пустыми ячейками
        xlsx_path = tmp_dir_path / "test_empty.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        
        sheet["A1"] = "Значение1"
        sheet["B1"] = None  # Пустая ячейка
        sheet["C1"] = "Значение2"
        sheet["A2"] = None
        sheet["B2"] = "Значение3"
        sheet["C2"] = None
        
        workbook.save(xlsx_path)
        
        # Конвертируем в csv
        csv_path = tmp_dir_path / "test_empty.csv"
        _convert_xlsx_to_csv(xlsx_path, csv_path)
        
        # Проверяем содержимое
        with csv_path.open("r", encoding="cp1251") as f:
            reader = csv.reader(f, delimiter=",")
            rows = list(reader)
            
            # Пустые ячейки должны быть заменены на пустые строки
            assert rows[0] == ["Значение1", "", "Значение2"], "Пустые ячейки должны быть пустыми строками"
            assert rows[1] == ["", "Значение3", ""], "Пустые ячейки должны быть пустыми строками"


def test_convert_empty_xlsx():
    """Проверяет обработку пустого xlsx файла."""
    with TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        
        # Создаем пустой xlsx файл (только с пустым листом)
        xlsx_path = tmp_dir_path / "test_empty_file.xlsx"
        workbook = Workbook()
        workbook.save(xlsx_path)
        
        # Конвертируем в csv
        csv_path = tmp_dir_path / "test_empty_file.csv"
        _convert_xlsx_to_csv(xlsx_path, csv_path)
        
        # Проверяем, что csv файл создан
        assert csv_path.exists(), "CSV файл должен быть создан даже для пустого xlsx"
        
        # Проверяем, что файл может быть прочитан (может быть пустым или содержать пустую строку)
        with csv_path.open("r", encoding="cp1251") as f:
            content = f.read()
            # Пустой файл или одна пустая строка - оба варианта допустимы для пустого листа


def test_convert_xlsx_to_csv_encoding():
    """Проверяет, что csv файл создается с кодировкой cp1251."""
    with TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        
        # Создаем xlsx файл с кириллицей
        xlsx_path = tmp_dir_path / "test_encoding.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Тест"
        sheet["B1"] = "Данные"
        sheet["A2"] = "Русский текст"
        
        workbook.save(xlsx_path)
        
        # Конвертируем в csv
        csv_path = tmp_dir_path / "test_encoding.csv"
        _convert_xlsx_to_csv(xlsx_path, csv_path)
        
        # Проверяем, что файл читается с кодировкой cp1251
        with csv_path.open("r", encoding="cp1251") as f:
            content = f.read()
            assert "Тест" in content, "Кириллица должна корректно читаться из файла с кодировкой cp1251"
            assert "Данные" in content, "Кириллица должна корректно читаться из файла с кодировкой cp1251"
            assert "Русский текст" in content, "Кириллица должна корректно читаться из файла с кодировкой cp1251"


def test_convert_xlsx_to_csv_delimiter():
    """Проверяет, что разделитель в csv файле - запятая."""
    with TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        
        # Создаем простой xlsx файл
        xlsx_path = tmp_dir_path / "test_delimiter.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = "Колонка1"
        sheet["B1"] = "Колонка2"
        sheet["C1"] = "Колонка3"
        
        workbook.save(xlsx_path)
        
        # Конвертируем в csv
        csv_path = tmp_dir_path / "test_delimiter.csv"
        _convert_xlsx_to_csv(xlsx_path, csv_path)
        
        # Проверяем, что разделитель - запятая
        with csv_path.open("r", encoding="cp1251") as f:
            line = f.readline().strip()
            # Разделитель должен быть запятой
            assert "," in line, "Разделитель должен быть запятой"
            # Проверяем, что значения разделены запятыми
            parts = line.split(",")
            assert len(parts) == 3, "Должно быть 3 колонки, разделенные запятыми"


import csv
from pathlib import Path
from tempfile import TemporaryDirectory

from aiogram import Router, F
from aiogram.types import Message, FSInputFile
from openpyxl import load_workbook

from services.utils import _get_random_quote, _is_allowed

router = Router()


def _convert_xlsx_to_csv(xlsx_path: Path, csv_path: Path):
    workbook = load_workbook(xlsx_path, data_only=True)
    sheet = workbook.active

    with csv_path.open("w", encoding="cp1251", newline="") as f:
        writer = csv.writer(f, delimiter=",")
        for row in sheet.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])


@router.message(F.document)
async def handle_document(message: Message):
    if not await _is_allowed(message.from_user.id):
        await message.answer("Привет, я тебя не узнал")
        return

    document = message.document
    if not document.file_name.lower().endswith(".xlsx"):
        await message.answer("Извини, но я работаю только с .xlsx файлами")
        return

    with TemporaryDirectory() as tmp_dir:
        tmp_dir_path = Path(tmp_dir)
        xlsx_path = tmp_dir_path / document.file_name
        # aiogram v3 Document не содержит метода download, используем бот
        await message.bot.download(document, destination=xlsx_path)

        csv_path = tmp_dir_path / f"{xlsx_path.stem}.csv"
        _convert_xlsx_to_csv(xlsx_path, csv_path)

        caption = await _get_random_quote()
        await message.answer_document(
            document=FSInputFile(csv_path),
            caption=caption,
        )
